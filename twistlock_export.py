#!/usr/bin/env python3
"""
twistlock_export.py

Convierte el CSV de vulnerabilidades exportado desde Prisma Cloud (Twistlock)
al formato de la Bitácora de Vulnerabilidades corporativa.

Genera un archivo .xlsx en una carpeta <nombre_input>-export/ con las columnas
de la bitácora en orden, listo para revisión y copy-paste.

Flujo de uso:
  1. Ejecutar el script sobre el CSV de Prisma -> genera el .xlsx.
  2. TRIAJE MANUAL del .xlsx antes de pegar: validar que cada versión vulnerable
     es real y se sufre en el código. Las que sean falso positivo se eliminan
     del .xlsx; así solo quedan las vulnerabilidades confirmadas.
  3. Copiar las filas confirmadas y pegarlas en la bitácora desde la columna A.
     Después, estirar (arrastrar) hacia abajo las fórmulas propias de la bitácora
     sobre las filas pegadas: la bitácora se autoajusta con los valores copiados.

Uso:
    python twistlock_export.py -i <fichero.csv>
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

# Forzar UTF-8 en stdout/stderr para evitar errores de codepage en Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

CURRENT_YEAR = date.today().year

# Columnas de la Bitácora de Vulnerabilidades corporativa, en orden exacto
# desde "ID" hasta "Resolution Date". El export replica esta estructura para
# permitir copy-paste directo: pega haciendo clic en la celda de la columna ID.
# Las columnas no mapeadas se generan vacías para respetar el alineamiento.
FIELDS = [
    "ID",
    "Hostname",
    "AB",
    "AP",
    "TAG",
    "IT Development Area",
    "COE",
    "State",
    "Service",
    "Origin",
    "Network",
    "Type",
    "Vulnerability Title",
    "Severity",
    "Domain",
    "Category ASVS",
    "ASVS ID",
    "OWASP Top 10",
    "PCI Status",
    "Threat Description",
    "Details",
    "Target",
    "Detection Date",
    "Countermeasure",
    "Environment",
    "Production Affected?",
    "References",
    "CVSS Base",
    "CVSS Score",
    "Easy of Exploit",
    "CVSS Version",
    "CVSS Vector",
    "Resolution Date",
    "Verificator",
    "Start Date",
    "Finish Date",
]

# Columnas A y B de la bitácora, anteriores a "ID". En el CSV/XLSX se anteponen
# vacías para que "ID" quede en la columna C y el copy-paste pueda hacerse desde
# la columna A de la bitácora. No forman parte de los datos (siempre vacías).
LEADING_COLS = ["", ""]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_cve_year(cve_id: str) -> int:
    m = re.search(r"CVE-(\d{4})-", cve_id, re.IGNORECASE)
    return int(m.group(1)) if m else 0


def is_cve(vuln_id: str) -> bool:
    """True si el identificador es un CVE (descarta GHSA, PRISMA, etc.)."""
    return vuln_id.strip().upper().startswith("CVE-")


# Orden de severidad para elegir el "CVE principal" del grupo
SEV_RANK = {"critical": 4, "high": 3, "medium": 2, "low": 1}


def sev_rank(row: dict) -> int:
    return SEV_RANK.get((row.get("Severity") or "").strip().lower(), 0)


def nvd_url(cve_id: str) -> str:
    """URL canónica de NVD a partir del CVE ID."""
    return f"https://nvd.nist.gov/vuln/detail/{cve_id.strip()}"


# Caracteres que, al inicio de una celda, Excel interpreta como fórmula.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def xlsx_safe(value):
    """Neutraliza la inyección de fórmulas en Excel (OWASP CSV Injection).

    Una celda cuyo texto empieza por '=', '+', '-', '@' o un carácter de control
    se ejecuta como fórmula al abrir el fichero en Excel. Los datos provienen de
    metadatos de imágenes de contenedor (nombre de paquete, descripción del CVE,
    Fix Status), potencialmente influenciables por un atacante, así que se
    antepone un apóstrofo para forzar que Excel los trate como texto literal.
    """
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def image_ref(row: dict) -> str:
    """Ruta legible de la imagen: 'Registry/Repository:Tag'.

    El campo 'Id' no sirve como identificador legible de forma uniforme: en los
    export de Prisma con scope 'registry' contiene la ruta completa de la imagen,
    pero en los de scope 'images' contiene el digest ('sha256:...'). Las columnas
    Registry/Repository/Tag están en ambos formatos, así que la ruta se
    reconstruye a partir de ellas. Si faltaran (export atípico), se cae al campo
    Id como último recurso.
    """
    registry = (row.get("Registry") or "").strip()
    repo = (row.get("Repository") or "").strip()
    tag = (row.get("Tag") or "").strip()
    if registry or repo:
        ref = "/".join(p for p in (registry, repo) if p)
        return f"{ref}:{tag}" if tag else ref
    return (row.get("Id") or "").strip()


def cvss_float(row: dict) -> float:
    """Lee la columna CVSS de forma tolerante (texto vacío o inválido -> 0.0)."""
    raw = (row.get("CVSS") or "").strip()
    try:
        return float(raw)
    except ValueError:
        return 0.0


def get_cvss_display(cvss_score: float, cve_id: str) -> str:
    """Devuelve el score CVSS o un texto explicativo cuando no está disponible.

    El número usa coma decimal (formato español de la bitácora): 7.8 -> "7,8".
    """
    if cvss_score > 0:
        return str(cvss_score).replace(".", ",")
    year = extract_cve_year(cve_id)
    if year >= CURRENT_YEAR:
        return "Pendiente de valoración NVD/NIST (CVE reciente)"
    return "Sin puntuación CVSS en NVD"


def format_fix_label(fix_status: str) -> str:
    """Convierte el Fix Status de Prisma en una etiqueta legible por CVE."""
    fs = (fix_status or "").strip()
    if fs.lower().startswith("fixed in"):
        return f"actualizar a {fs[len('fixed in'):].strip()}"
    if fs.lower() == "deferred":
        return "sin parche disponible"
    if fs.lower() in ("needed", "open"):
        return "parche pendiente"
    return "sin información de fix"


def parse_csv(input_path: Path):
    """
    Lee el CSV de Prisma Cloud. Devuelve (filas_no_OS, num_filas_OS).
    El conteo de OS se hace aquí sobre registros CSV reales (no líneas físicas)
    para ser robusto frente a descripciones con saltos de línea embebidos.
    """
    rows = []
    os_count = 0
    with open(input_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Type", "").strip() == "OS":
                os_count += 1
                continue
            rows.append(row)
    return rows, os_count


def group_and_build(rows: list) -> list:
    """
    Agrupa por (imagen, paquete, versión) y construye una fila de bitácora por
    grupo. Los CVEs se deduplican preservando orden de aparición y se descartan
    identificadores no-CVE (GHSA, PRISMA); en Details cada CVE lleva su versión
    de fix (Fix Status). El CVE principal del grupo (mayor severidad; a igualdad,
    mayor CVSS) aporta Threat Description, CVSS Base/Score y References.
    Countermeasure es un mensaje genérico fijo. Severity se deja vacío: la
    bitácora lo autocalcula desde el CVSS.

    Orden de salida en dos bloques: primero las entradas con score CVSS (de mayor
    a menor) y después las que no tienen score, por severidad descendente; como
    desempate, por nombre de paquete.
    """
    groups = defaultdict(list)
    for row in rows:
        key = (row["Id"], row["Packages"], row["Package Version"])
        groups[key].append(row)

    entries = []  # (no_score, sev_rank, cvss_val, pkg, row) para ordenar al final
    for (_, pkg, version), vulns in groups.items():
        # CVE IDs deduplicados, solo CVE-*, preservando orden de aparición.
        # cve_fix: CVE ID → etiqueta de fix (primera aparición del CVE gana).
        seen = set()
        cve_ids = []
        cve_fix = {}
        for v in vulns:
            vid = v["CVE ID"].strip()
            if is_cve(vid) and vid not in seen:
                seen.add(vid)
                cve_ids.append(vid)
                cve_fix[vid] = format_fix_label(v.get("Fix Status", ""))

        # Pool para elegir el "CVE principal": solo CVEs si los hay.
        # Fallback defensivo: si el grupo no tuviera ningún CVE, usar todos.
        cve_vulns = [v for v in vulns if is_cve(v["CVE ID"])]
        pool = cve_vulns if cve_vulns else vulns
        # Principal = mayor severidad; a igualdad de severidad, mayor CVSS.
        # Así Severity, CVSS, Threat Description y References salen del mismo
        # CVE y reflejan el peor caso del grupo de forma coherente.
        top = sorted(pool, key=lambda v: (sev_rank(v), cvss_float(v)), reverse=True)[0]
        top_cve = top["CVE ID"].strip()

        # Si no quedó ningún CVE (caso límite), no perder los identificadores
        ids_display = cve_ids or list(dict.fromkeys(v["CVE ID"].strip() for v in vulns))

        title = (
            f"{ids_display[0]} en {pkg} ({version})"
            if len(ids_display) == 1
            else f"Múltiples CVEs en {pkg} ({version})"
        )

        # Threat Description: descripción del CVE de mayor criticidad del grupo.
        threat = (top.get("Description", "") or "").strip()

        cvss_display = get_cvss_display(cvss_float(top), top_cve)
        references = nvd_url(top_cve) if is_cve(top_cve) else top.get("Vulnerability Link", "").strip()

        # Ruta legible de la imagen (Registry/Repository:Tag). Todas las filas
        # del grupo comparten imagen (misma clave 'Id'), así que vale cualquiera.
        container_ref = image_ref(vulns[0])

        # Valores fijos para todos los exports de este proyecto:
        # State=Open, Type=Application, Domain=Configuration Error,
        # ASVS ID=ASVS-14.2.1. Si se reutilizara el script en otro proyecto con
        # otros valores, habría que parametrizarlos aquí.
        row = dict.fromkeys(FIELDS, "")
        row.update({
            "Hostname": container_ref,
            "State": "Open",
            "Type": "Application",
            "Vulnerability Title": title,
            # Severity se deja vacío: la bitácora lo autocompleta a partir del CVSS.
            "Domain": "Configuration Error",
            "ASVS ID": "ASVS-14.2.1",
            "Threat Description": threat,
            "Details": (
                f"La versión {version} de {pkg} tiene los siguientes CVEs afectados:\n"
                + "\n".join(f"{cid} → {cve_fix[cid]}" for cid in cve_ids)
                if cve_ids
                else f"La versión {version} de {pkg} tiene los siguientes IDs afectados: {', '.join(ids_display)}."
            ),
            "Target": container_ref,
            "Countermeasure": (
                f"Se recomienda actualizar {pkg} a la última versión disponible "
                "del proveedor y revisar las versiones de corrección indicadas en Details."
            ),
            "References": references,
            "CVSS Base": cvss_display,
            "CVSS Score": cvss_display,
        })

        # Dos bloques: scored (CVSS > 0) primero, sin-score después.
        # Dentro de cada bloque: sev_rank desc, CVSS desc, pkg asc.
        cvss_val = cvss_float(top)
        no_score = 0 if cvss_val > 0 else 1
        entries.append((no_score, sev_rank(top), cvss_val, pkg.lower(), row))

    entries.sort(key=lambda e: (e[0], -e[1], -e[2], e[3]))
    return [row for _, _, _, _, row in entries]


# ---------------------------------------------------------------------------
# Exporters
# ---------------------------------------------------------------------------

def export_xlsx(rows: list, output_path: Path) -> None:
    """
    Excel con las columnas de la bitácora en orden, cabeceras coloreadas y texto
    ajustado. Es el único formato de salida: pensado para el triaje manual previo
    (revisar/eliminar falsos positivos) y el posterior copy-paste a la bitácora.

    Antepone 2 columnas vacías (A y B de la bitácora) para que 'ID' caiga en la
    columna C y el copy-paste pueda hacerse desde la columna A.
    """
    if not XLSX_AVAILABLE:
        print("ERROR: falta la dependencia 'openpyxl', necesaria para generar el .xlsx.")
        print("       Instálala con: pip install -r requirements.txt")
        raise SystemExit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora Export"

    # Estilos
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Columnas en orden de la bitácora: las 2 iniciales (A/B) + FIELDS.
    # Offset de 2 para que "ID" quede en la columna C, igual que la bitácora.
    all_cols = LEADING_COLS + FIELDS

    # Cabecera
    for col, field in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 22

    # Datos (las 2 columnas iniciales quedan vacías)
    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = LEADING_COLS + [row.get(field, "") for field in FIELDS]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=xlsx_safe(value))
            cell.alignment = wrap_top
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 70

    # Anchos de columna (las columnas vacías mantienen un ancho discreto)
    col_widths = {
        "": 4,
        "ID": 10, "Hostname": 50, "AB": 10, "AP": 14, "TAG": 22,
        "IT Development Area": 20, "COE": 18, "State": 8, "Service": 14,
        "Origin": 18, "Network": 12, "Type": 12, "Vulnerability Title": 42,
        "Severity": 10, "Domain": 20, "Category ASVS": 14, "ASVS ID": 14,
        "OWASP Top 10": 22, "PCI Status": 10, "Threat Description": 55,
        "Details": 40, "Target": 50, "Detection Date": 14, "Countermeasure": 42,
        "Environment": 14, "Production Affected?": 16, "References": 40,
        "CVSS Base": 16, "CVSS Score": 16, "Easy of Exploit": 14,
        "CVSS Version": 12, "CVSS Vector": 20, "Resolution Date": 14,
        "Verificator": 14, "Start Date": 14, "Finish Date": 14,
    }
    for col, field in enumerate(all_cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = col_widths.get(field, 14)

    # Congelar cabecera y las 3 primeras columnas (A, B, ID) al hacer scroll
    ws.freeze_panes = "D2"
    wb.save(output_path)
    print(f"  [XLSX] {output_path.name}")



# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        prog="twistlock_export.py",
        description="Convierte el CSV de Prisma Cloud al formato de la Bitacora de Vulnerabilidades.",
        epilog="""\
Flujo:
  1. Ejecutar sobre el CSV de Prisma -> genera el .xlsx
  2. Triaje manual: eliminar falsos positivos del .xlsx
  3. Pegar filas confirmadas en la bitacora desde columna A y estirar formulas

Campos autogenerados: Hostname, State, Type, Vulnerability Title, Domain,
  ASVS ID, Threat Description, Details, Target, Countermeasure, References,
  CVSS Base, CVSS Score. Severity se deja vacio (formula de la bitacora).

Ejemplos:
  twistlock_export.py -i vulns.csv
  twistlock_export.py -i vulns.csv -o resultado.xlsx
  twistlock_export.py -i vulns.csv -o C:\\exports\\proyecto.xlsx
""",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "-i", "--input", required=True, metavar="CSV",
        help="CSV exportado desde Prisma Cloud / Twistlock",
    )
    parser.add_argument(
        "-o", "--output", metavar="FICHERO.xlsx",
        help="Ruta del .xlsx de salida (incluir extension). "
             "Por defecto: <dir_csv>/<nombre_csv>-export/<nombre_csv>.xlsx",
    )
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"ERROR: archivo no encontrado — {input_path}")
        raise SystemExit(1)

    if args.output:
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = input_path.parent / output_path
        output_path = output_path.resolve()
    else:
        output_dir = input_path.parent / f"{input_path.stem}-export"
        output_path = output_dir / f"{input_path.stem}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\nInput : {input_path.name}")
    print(f"Output: {output_path}\n")

    rows_raw, os_count = parse_csv(input_path)
    bitacora_rows = group_and_build(rows_raw)

    print(f"Filas en el CSV       : {len(rows_raw) + os_count}")
    print(f"Excluidas (OS)        : {os_count}")
    print(f"Procesadas            : {len(rows_raw)}")
    print(f"Entradas en bitacora  : {len(bitacora_rows)} (agrupadas por paquete)\n")

    try:
        export_xlsx(bitacora_rows, output_path)
    except PermissionError as e:
        print(f"\nERROR: no se pudo escribir '{e.filename}'.")
        print("       Probablemente lo tienes abierto en Excel. Ciérralo y reejecuta.")
        raise SystemExit(1)

    print(f"\nExport completado en: {output_path.parent}")


if __name__ == "__main__":
    main()
